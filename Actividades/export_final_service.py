import os
import pandas as pd
import openpyxl
import copy
from datetime import datetime
from config import TEMPLATE_INFORME_FINAL, logger
import unicodedata
from utils import medir_tiempo

@medir_tiempo
def generar_informe_final_resumen(df, output_path, contrato_data=None, usuario=None):
    """
    Versión v6.7: Genera el Informe Final con filtro por usuario y precisión absoluta.
    df: DataFrame con todos los registros (será filtrado por usuario)
    contrato_data: dict con info del contrato
    usuario: El usuario dueño del reporte (para filtrar actividades admin)
    """
    try:
        import openpyxl
        from config import TEMPLATE_INFORME_FINAL, logger

        logger.info(f"Generando Informe Final Concentrado. Registros iniciales: {len(df)}")

        if df.empty:
            logger.warning("Generar Informe Final: DataFrame vacío")
            return False

        # --- FILTRO POR USUARIO (v6.7) ---
        if usuario:
            # Si el usuario NO es admin, solo ver sus registros.
            # Si es admin, puede ver todo o se filtra por un usuario específico pasado aquí.
            df = df[df['USUARIO'] == usuario].copy()
            logger.info(f"Reporte Final: Filtrado para usuario '{usuario}'. Registros: {len(df)}")
        
        if df.empty:
            logger.warning(f"Reporte Final: No hay registros para el usuario '{usuario}' después de filtrar.")
            return False

        if not os.path.exists(TEMPLATE_INFORME_FINAL):
            logger.error(f"Plantilla no encontrada: {TEMPLATE_INFORME_FINAL}")
            return False

        # Obtener datos de contrato
        if contrato_data is None:
            try:
                from database import obtener_configuracion_usuario
                cfg = obtener_configuracion_usuario('admin') or {}
                contrato_data = cfg.get('datos_contrato', {}) if isinstance(cfg, dict) else {}
            except Exception:
                contrato_data = {}

        wb = openpyxl.load_workbook(TEMPLATE_INFORME_FINAL)
        ws = wb.active
        
        # 1. Agrupar y contar actividades
        if 'TIPO DE ACTIVIDAD' in df.columns:
            resumen = df['TIPO DE ACTIVIDAD'].value_counts().reset_index()
            resumen.columns = ['Actividad', 'Cantidad']
        else:
            resumen = pd.DataFrame(columns=['Actividad', 'Cantidad'])

        # 2. Reemplazo de placeholders y Smart Fill (v6 Definitive)
        def _get_primary_cell(sheet, cell):
            for m_range in sheet.merged_cells.ranges:
                if cell.coordinate in m_range:
                    return sheet.cell(row=m_range.min_row, column=m_range.min_col)
            return cell

        def _build_contrato_values(df, contrato_data, usuario=None):
            vals = {}
            n_ = lambda x: (contrato_data.get(x) or '').upper() if contrato_data else ''
            vals['NRO_CONTRATO'] = n_('nro')
            vals['OBJETO'] = n_('objeto')
            vals['NOMBRE_CONTRATISTA'] = n_('nombre') or (usuario.upper() if usuario else '')
            vals['CEDULA'] = n_('cedula')
            vals['SUPERVISOR'] = n_('supervisor')
            
            # Fechas (v6.6)
            hoy = datetime.now()
            vals['FECHA_HOY'] = hoy.strftime('%d/%m/%Y')
            
            rango = 'N/A'
            try:
                if not df.empty and 'FECHA' in df.columns:
                    fechas_dt = pd.to_datetime(df['FECHA'], errors='coerce').dropna()
                    if not fechas_dt.empty:
                        rango = f"{fechas_dt.min().strftime('%d/%m/%Y')} al {fechas_dt.max().strftime('%d/%m/%Y')}"
            except Exception: pass
            vals['RANGO_FECHAS'] = rango
            return vals

        contrato_values = _build_contrato_values(df, contrato_data, usuario)
        
        def _normalize(s):
            s = str(s).strip().upper().replace(':','')
            # Eliminar tildes y caracteres especiales
            s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
            return ''.join(s.split())

        # Mapa de etiquetas preciso y normalizado
        label_map_norm = {
            "NRO_CONTRATO": ["NOCONVENIOOCONTRATO", "N°CONVENIOOCONTRATO", "NROCONVENIOOCONTRATO", "NOCONTRATO", "NROCONTRATO", "NºCONTRATO", "CONVENIO", "CONTRATO"],
            "OBJETO": ["OBJETODELCONTRATO", "OBJETO"],
            "NOMBRE_CONTRATISTA": ["NOMBREDELCONTRATISTA", "NOMBRECONTRATISTA", "CONTRATISTA", "ELABORADOPOR"],
            "CEDULA": ["NODEIDENTIFICACION", "CEDULA", "NIT", "IDENTIFICACION", "CC"], 
            "RANGO_FECHAS": ["FECHADEACTIVIDADES", "RANGODEFECHAS"],
            "FECHA_HOY": ["FECHADEINFORME", "FECHADEINFOR"]
        }

        filled_coords = set()
        
        # Fase 1: Placeholders
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=1, max_col=ws.max_column):
            for cell in row:
                val = str(cell.value) if cell.value else ""
                if '{{' in val and '}}' in val:
                    primary = _get_primary_cell(ws, cell)
                    new_val = str(primary.value)
                    for k, v in contrato_values.items():
                        if '{{' + k + '}}' in new_val:
                            new_val = new_val.replace('{{' + k + '}}', v)
                            filled_coords.add(primary.coordinate)
                            logger.info(f"PH REPLACEMENT: {k} en {primary.coordinate}")
                    primary.value = new_val

        # Fase 2: Smart Fill (v6.6 Aggressive Date Overwrite)
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=1, max_col=ws.max_column):
            for cell in row:
                if not cell.value or isinstance(cell.value, (int, float)): continue
                val_norm = _normalize(cell.value)
                
                for key, labels in label_map_norm.items():
                    if any(lbl in val_norm for lbl in labels):
                        label_primary = _get_primary_cell(ws, cell)
                        for c_off in range(1, 6):
                            neighbor = ws.cell(row=cell.row, column=cell.column + c_off)
                            neighbor_primary = _get_primary_cell(ws, neighbor)
                            
                            if neighbor_primary.coordinate != label_primary.coordinate:
                                n_val = str(neighbor_primary.value or "").upper()
                                if neighbor_primary.coordinate not in filled_coords:
                                    # Overwrite keywords for dates (v6.6)
                                    meses = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", 
                                            "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
                                    keywords = ["NOMBRE", "---", "FIRMA", "IDENTIFICA", "JOHNATAN", "FELIPE", 
                                               "ALFONSO", "2025", "2024", "2023", "CONTRATO", "PRESTACION", "SERVICIOS", "1040"]
                                    keywords.extend(meses)
                                    
                                    if not n_val or any(x in n_val for x in keywords):
                                        neighbor_primary.value = contrato_values[key]
                                        filled_coords.add(neighbor_primary.coordinate)
                                        logger.info(f"SMART FILL v6.6: {key} en {neighbor_primary.coordinate} (vía label '{val_norm}')")
                                        break
                        if any(c in filled_coords for c in [ws.cell(row=cell.row, column=cell.column + i).coordinate for i in range(1, 6)]):
                            break

        # 3. Guardar diseño de tabla (Fila 7)
        header_styles = []
        row7_merges = []
        for col in range(1, 11):
            c = ws.cell(row=7, column=col)
            if hasattr(c, '_style'):
                header_styles.append(c._style)
            else:
                header_styles.append(None)
        for m_range in list(ws.merged_cells.ranges):
            if m_range.min_row == 7 and m_range.max_row == 7:
                row7_merges.append((m_range.min_col, m_range.max_col))
        if not row7_merges: row7_merges = [(1, 7), (8, 10)]

        # 4. Limpiar TODA la zona a partir de fila 8
        for m_range in list(ws.merged_cells.ranges):
            if m_range.min_row >= 8:
                try: ws.unmerge_cells(str(m_range))
                except Exception: pass

        # 5. Insertar filas
        rows_to_insert = len(resumen)
        start_row = 8
        if rows_to_insert > 1:
            ws.insert_rows(start_row, amount=rows_to_insert - 1)

        # 6. Escribir Datos y Replicar Estructura EXACTA
        current_row = 8
        for _, r in resumen.iterrows():
            ws.cell(row=current_row, column=1, value=r['Actividad'])
            ws.cell(row=current_row, column=8, value=r['Cantidad'])
            for min_c, max_c in row7_merges:
                if min_c != max_c:
                    try: ws.merge_cells(start_row=current_row, start_column=min_c, end_row=current_row, end_column=max_c)
                    except Exception: pass
            for c in range(1, 11):
                cell = ws.cell(row=current_row, column=c)
                style_id = header_styles[c-1]
                if style_id is not None:
                    cell._style = style_id
                if c >= 8: cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            current_row += 1
            
        # 7. Total General
        ws.cell(row=current_row, column=1, value="TOTAL GENERAL DE ACTIVIDADES:")
        ws.cell(row=current_row, column=8, value=resumen['Cantidad'].sum())
        for min_c, max_c in row7_merges:
            if min_c != max_c:
                try: ws.merge_cells(start_row=current_row, start_column=min_c, end_row=current_row, end_column=max_c)
                except Exception: pass
        for c in range(1, 11):
            cell = ws.cell(row=current_row, column=c)
            style_id = header_styles[c-1]
            if style_id is not None:
                cell._style = style_id
            if c >= 8: cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
        ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_row, column=8).font = openpyxl.styles.Font(bold=True)

        wb.save(output_path)
        return True
    except Exception as e:
        logger.error(f"Error generando informe final: {e}")
        return False
