"""
Servicio de exportación y generación de reportes.
Separado de database.py para responsabilidad única.
"""

import os
import copy
import pandas as pd
import unicodedata
from datetime import datetime
from config import TEMPLATE_EXCEL, logger
from database import cargar_registros
from utils import medir_tiempo


@medir_tiempo
def exportar_registros_filtrados(fecha_inicio=None, fecha_fin=None, usuario=None, actividad=None):
    """Exporta registros filtrados. Retorna (DataFrame, dict_estadísticas)"""
    try:
        df = cargar_registros(usuario)
        if df.empty:
            return pd.DataFrame(), {}
        
        # Parsear fechas
        if 'FECHA' in df.columns:
            df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
        
        # Aplicar filtros
        if fecha_inicio:
            df = df[df['FECHA'] >= pd.to_datetime(fecha_inicio)]
        if fecha_fin:
            df = df[df['FECHA'] <= pd.to_datetime(fecha_fin)]
        if actividad and actividad != 'Todas' and 'TIPO DE ACTIVIDAD' in df.columns:
            df = df[df['TIPO DE ACTIVIDAD'] == actividad]
        
        # Agrupamiento y ordenamiento solicitado
        if not df.empty:
            sort_cols = []
            if 'TIPO DE ACTIVIDAD' in df.columns: sort_cols.append('TIPO DE ACTIVIDAD')
            if 'FECHA' in df.columns: sort_cols.append('FECHA')
            if sort_cols:
                df = df.sort_values(by=sort_cols)
        
        stats = _calcular_estadisticas(df)
        return df, stats
    except Exception as e:
        logger.error(f"Error exportando registros: {e}")
        return pd.DataFrame(), {}


def _calcular_estadisticas(df):
    """Calcula estadísticas básicas de un DataFrame"""
    if df.empty:
        return {}
    
    stats = {
        'total_registros': len(df),
        'fecha_inicio': _format_fecha(df, 'min'),
        'fecha_fin': _format_fecha(df, 'max'),
    }
    
    # Conteos por columna
    for col, key in [('TIPO DE ACTIVIDAD', 'conteo_por_actividad'),
                     ('TIPO DE SOLICITUD', 'conteo_por_solicitud'),
                     ('MEDIO DE SOLICITUD', 'conteo_por_medio')]:
        if col in df.columns:
            counts = df[col].value_counts().to_dict()
            stats[key] = {str(k): int(v) for k, v in counts.items()}
    
    return stats


def _format_fecha(df, func):
    """Formatea fecha min/max de un DataFrame"""
    if 'FECHA' not in df.columns or df['FECHA'].empty:
        return 'N/A'
    try:
        val = getattr(df['FECHA'], func)()
        return val.strftime('%Y-%m-%d') if pd.notna(val) else 'N/A'
    except Exception:
        return 'N/A'


@medir_tiempo
def obtener_estadisticas_exportacion(usuario=None, fecha_inicio=None, fecha_fin=None):
    """Obtiene estadísticas generales y datos para gráficos con soporte de filtros"""
    empty_result = {
        'fecha_min': 'N/A', 'fecha_max': 'N/A',
        'total_registros': 0, 'total_tipos_actividad': 0,
        'ultima_exportacion': 'Nunca',
        'chart_actividades': {'labels': [], 'data': []},
        'chart_cumplimiento': {'labels': [], 'data': []},
        'chart_linea': {'labels': [], 'data': []},
        'usuarios': []
    }
    
    try:
        df, _ = exportar_registros_filtrados(
            fecha_inicio=fecha_inicio, 
            fecha_fin=fecha_fin, 
            usuario=usuario
        )
        if df.empty:
            return empty_result
        
        # Parseo de fechas (exportar_registros_filtrados ya hace parte del proceso)
        if 'FECHA' in df.columns:
            df['FECHA_DT'] = pd.to_datetime(df['FECHA'], errors='coerce')
            df = df.dropna(subset=['FECHA_DT'])
        
        if df.empty:
            return empty_result
        
        # Gráfico: Actividades (Todas, según petición del usuario)
        counts_act = df['TIPO DE ACTIVIDAD'].value_counts()
        chart_actividades = {
            'labels': counts_act.index.tolist(),
            'data': counts_act.values.tolist()
        }
        
        # Gráfico: Cumplimiento
        cumplimiento = df['CUMPLIDO'].value_counts() if 'CUMPLIDO' in df.columns else pd.Series()
        chart_cumplimiento = {
            'labels': cumplimiento.index.tolist(),
            'data': cumplimiento.values.tolist()
        }
        
        # Gráfico: Línea temporal
        df_sorted = df.sort_values('FECHA_DT')
        linea = df_sorted['FECHA_DT'].dt.date.value_counts().sort_index()
        # Si hay demasiados días, mostrar los últimos 90 para no saturar
        if len(linea) > 90:
            linea = linea.tail(90)
            
        chart_linea = {
            'labels': [d.strftime('%d/%m') for d in linea.index],
            'data': linea.values.tolist()
        }
        
        # Estadística por usuario
        user_stats = []
        if 'USUARIO' in df.columns:
            for user, group in df.groupby('USUARIO'):
                total_user = len(group)
                cumplidos = len(group[group['CUMPLIDO'] == 'Sí'])
                porcentaje = f"{(cumplidos/total_user)*100:.1f}%" if total_user > 0 else "0%"
                ultima = group['FECHA_DT'].max().strftime('%Y-%m-%d %H:%M') if not group.empty else "N/A"
                user_stats.append({
                    'usuario': user,
                    'total': total_user,
                    'cumplimiento': porcentaje,
                    'ultima': ultima
                })
        
        return {
            'fecha_min': df['FECHA_DT'].min().strftime('%Y-%m-%d'),
            'fecha_max': df['FECHA_DT'].max().strftime('%Y-%m-%d'),
            'total_registros': len(df),
            'total_tipos_actividad': df['TIPO DE ACTIVIDAD'].nunique() if 'TIPO DE ACTIVIDAD' in df.columns else 0,
            'ultima_exportacion': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'chart_actividades': chart_actividades,
            'chart_cumplimiento': chart_cumplimiento,
            'chart_linea': chart_linea,
            'usuarios': user_stats
        }
    except Exception as e:
        logger.error(f"Error obteniendo estadísticas: {e}")
        return empty_result


@medir_tiempo
def generar_reporte_excel(df, estadisticas, output_path):
    """Genera un archivo Excel con datos + estadísticas"""
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Registros', index=False)
            
            stats_df = pd.DataFrame([
                ['Total de registros', estadisticas.get('total_registros', 0)],
                ['Fecha inicio', estadisticas.get('fecha_inicio', 'N/A')],
                ['Fecha fin', estadisticas.get('fecha_fin', 'N/A')]
            ], columns=['Métrica', 'Valor'])
            stats_df.to_excel(writer, sheet_name='Estadísticas', index=False)
            
            if 'conteo_por_actividad' in estadisticas:
                act_df = pd.DataFrame(
                    estadisticas['conteo_por_actividad'].items(),
                    columns=['Actividad', 'Cantidad']
                )
                act_df.to_excel(writer, sheet_name='Estadísticas', startrow=5, index=False)
        
        return True
    except Exception as e:
        logger.error(f"Error generando reporte Excel: {e}")
        return False


@medir_tiempo
def generar_informe_template(df, output_path, contrato_data=None):
    """Genera informe usando la plantilla Excel institucional"""
    try:
        import openpyxl
        from config import TEMPLATE_EXCEL
        # Si no se pasaron datos de contrato, intentar obtenerlos de la
        # configuración global guardada en la base de datos bajo 'admin'.
        if contrato_data is None:
            try:
                from database import obtener_configuracion_usuario
                cfg = obtener_configuracion_usuario('admin') or {}
                contrato_data = cfg.get('datos_contrato', {}) if isinstance(cfg, dict) else {}
            except Exception:
                contrato_data = {}
        from config import logger, TEMPLATE_EXCEL, DATA_DIR, BASE_DIR
        template_path = TEMPLATE_EXCEL
        if not os.path.exists(template_path):
            candidates = [
                os.path.join(DATA_DIR, "INFORME DE ACTIVIDADES - copia.xlsx"),
                os.path.join(BASE_DIR, "INFORME DE ACTIVIDADES - copia.xlsx"),
                os.path.join(os.getcwd(), "INFORME DE ACTIVIDADES - copia.xlsx"),
                os.path.join(os.path.dirname(BASE_DIR), "INFORME DE ACTIVIDADES - copia.xlsx"),
                os.path.join(os.path.dirname(os.getcwd()), "INFORME DE ACTIVIDADES - copia.xlsx"),
            ]
            for cand in candidates:
                if os.path.exists(cand):
                    template_path = cand
                    logger.info(f"Plantilla encontrada por búsqueda alternativa: {template_path}")
                    break
        if not os.path.exists(template_path):
            logger.error(f"Plantilla no encontrada - Se intentó cargar desde: {TEMPLATE_EXCEL}")
            # Fallback: generar reporte básico sin plantilla
            try:
                stats = _calcular_estadisticas(df)
                return generar_reporte_excel(df, stats, output_path)
            except Exception:
                return False
            
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            logger.info("Plantilla cargada correctamente")
        except Exception as e:
            logger.error(f"Error al cargar el libro de Excel: {e}")
            return False
        # Reemplazo de placeholders en la plantilla para mayor flexibilidad.
        # Soportamos: {{NRO_CONTRATO}}, {{OBJETO}}, {{NOMBRE_CONTRATISTA}}, {{CEDULA}}, {{SUPERVISOR}}, {{RANGO_FECHAS}}
        def _build_contrato_values(df, contrato_data):
            vals = {}
            nro = (contrato_data.get('nro') if contrato_data else '') or ''
            objeto = (contrato_data.get('objeto') if contrato_data else '') or ''
            nombre = (contrato_data.get('nombre') if contrato_data else '') or ''
            cedula = (contrato_data.get('cedula') if contrato_data else '') or ''
            supervisor = (contrato_data.get('supervisor') if contrato_data else '') or ''

            # Si no tenemos nombre en contrato_data, intentar obtenerlo de columnas del DataFrame
            if not nombre:
                candidate_cols = [
                    'NOMBRE CONTRATISTA', 'CONTRATISTA', 'NOMBRE',
                    'NOMBRE_COMPLETO', 'NOMBRE_USUARIO', 'NOMBRE DEL CONTRATISTA'
                ]
                found = None
                for c in candidate_cols:
                    if c in df.columns:
                        found = c
                        break
                if found:
                    vals_list = df[found].dropna().unique().tolist()
                    if len(vals_list) == 1:
                        nombre = str(vals_list[0])
                    elif len(vals_list) > 1:
                        preview = [str(v).strip() for v in vals_list[:3]]
                        nombre = ', '.join(preview) + ('...' if len(vals_list) > 3 else '')
                else:
                    if 'USUARIO' in df.columns:
                        usuarios = df['USUARIO'].dropna().unique().tolist()
                        if len(usuarios) == 1:
                            nombre = str(usuarios[0])
                        else:
                            nombre = 'VARIOS'

            # Rango de fechas
            rango = ''
            try:
                if not df.empty and 'FECHA' in df.columns:
                    fechas_dt = pd.to_datetime(df['FECHA'], errors='coerce').dropna()
                    if not fechas_dt.empty:
                        rango = f"{fechas_dt.min().strftime('%d/%m/%Y')} al {fechas_dt.max().strftime('%d/%m/%Y')}"
            except Exception:
                rango = ''

            vals['NRO_CONTRATO'] = str(nro).upper() if nro else ''
            vals['OBJETO'] = str(objeto).upper() if objeto else ''
            vals['NOMBRE_CONTRATISTA'] = str(nombre).upper() if nombre else 'N/D'
            vals['CEDULA'] = str(cedula) if cedula else ''
            vals['SUPERVISOR'] = str(supervisor).upper() if supervisor else ''
            vals['RANGO_FECHAS'] = rango
            return vals

        def _replace_placeholders_in_sheet(sheet, values_map):
            # Iterar por todo el rango usado de la hoja
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and '{{' in cell.value and '}}' in cell.value:
                        new_val = cell.value
                        for k, v in values_map.items():
                            placeholder = '{{' + k + '}}'
                            if placeholder in new_val:
                                new_val = new_val.replace(placeholder, v)
                        cell.value = new_val

        contrato_values = _build_contrato_values(df, contrato_data)
        # Reemplazar en todas las hojas por seguridad; si no hay placeholders
        # presentes hacemos fallback a las celdas fijas para mantener
        # compatibilidad con plantillas antiguas.
        replaced_any = False
        for sht in wb.worksheets:
            # marcar antes del cambio para detectar si se reemplazó algo
            before_has = False
            for row in sht.iter_rows(min_row=1, max_row=sht.max_row, min_col=1, max_col=sht.max_column):
                for cell in row:
                    if isinstance(cell.value, str) and '{{' in cell.value and '}}' in cell.value:
                        before_has = True
                        break
                if before_has:
                    break

            if before_has:
                _replace_placeholders_in_sheet(sht, contrato_values)
                replaced_any = True

        # Completar cabecera por etiquetas visibles (compatibilidad con plantillas sin placeholders)
        def _normalize(s):
            s = str(s).strip().upper().replace(':','')
            s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
            return ''.join(s.split())
        label_map = {
            'NRO_CONTRATO': [
                "NOCONVENIOOCONTRATO", "N°CONVENIOOCONTRATO", "NROCONVENIOOCONTRATO",
                "NOCONTRATO", "NROCONTRATO", "NºCONTRATO", "CONVENIO", "CONTRATO"
            ],
            'OBJETO': ["OBJETODELCONTRATO", "OBJETO"],
            'NOMBRE_CONTRATISTA': ["NOMBREDELCONTRATISTA", "NOMBRECONTRATISTA", "CONTRATISTA", "ELABORADOPOR"],
            'CEDULA': ["NODEIDENTIFICACION", "CEDULA", "NIT", "IDENTIFICACION", "CC"],
            'RANGO_FECHAS': ["FECHADEACTIVIDADES", "RANGODEFECHAS", "FECHADEINFORME", "FECHADEINFOR"]
        }
        def _get_primary_cell(sheet, cell):
            for m_range in sheet.merged_cells.ranges:
                if cell.coordinate in m_range:
                    return sheet.cell(row=m_range.min_row, column=m_range.min_col)
            return cell

        def _write_next_to_label(sheet, label_keys, value):
            if not value:
                return False
            max_r = min(sheet.max_row, 30)
            max_c = min(sheet.max_column, 15)
            for r in range(1, max_r+1):
                for c in range(1, max_c+1):
                    cell = sheet.cell(row=r, column=c)
                    if isinstance(cell.value, str):
                        norm = _normalize(cell.value)
                        for key in label_keys:
                            if key in norm:
                                # Encontrar la primera celda a la derecha que no sea la misma (por celdas combinadas)
                                label_primary = _get_primary_cell(sheet, cell)
                                for c_off in range(1, 6):
                                    neighbor = sheet.cell(row=r, column=c + c_off)
                                    neighbor_primary = _get_primary_cell(sheet, neighbor)
                                    if neighbor_primary.coordinate != label_primary.coordinate:
                                        try:
                                            neighbor_primary.value = value
                                            return True
                                        except Exception:
                                            return False
            return False

        # Intentar escribir en la hoja activa
        try:
            for key, keys_list in label_map.items():
                _write_next_to_label(ws, keys_list, contrato_values.get(key, ''))
        except Exception as e:
            logger.error(f"Error in _write_next_to_label: {e}")
            pass

 

        # Rango de fechas (Fila 6) - Mantener si es necesario o mover al final
        if not df.empty and 'FECHA' in df.columns:
            fechas_dt = pd.to_datetime(df['FECHA'], errors='coerce').dropna()
            if not fechas_dt.empty:
                ws.cell(row=6, column=3,
                        value=f"{fechas_dt.min().strftime('%d/%m/%Y')} al {fechas_dt.max().strftime('%d/%m/%Y')}")
                ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=10) # Unificar ancho de fechas
        
        # Guardar estilos base de fila 8
        base_styles = []
        for col in range(1, 10):
            cell = ws.cell(row=8, column=col)
            if hasattr(cell, '_style'):
                base_styles.append(cell._style)
            else:
                base_styles.append(None)
        
        # Diccionario de meses en español
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        ahora = datetime.now()
        mes_actual = meses[ahora.month - 1]
        anio_actual = ahora.year

        # ------------------------------------------------------------------
        # FIX CORRUPCIÓN EXCEL: Primero descombinamos celdas existentes 
        # en la zona de trabajo para evitar conflictos al insertar/combinar.
        # ------------------------------------------------------------------
        merged_ranges = list(ws.merged_cells.ranges)
        for m_range in merged_ranges:
            if m_range.min_row >= 7:
                try:
                    ws.unmerge_cells(str(m_range))
                except Exception: pass

        # Calcular cuántas filas necesitamos exactamente
        df_reporte = df.copy()
        if 'TIPO DE ACTIVIDAD' in df_reporte.columns:
            df_reporte = df_reporte.sort_values(by=['TIPO DE ACTIVIDAD', 'FECHA'])
            resumen_counts = df_reporte['TIPO DE ACTIVIDAD'].value_counts().to_dict()
            num_grupos = len(resumen_counts)
        else:
            resumen_counts = {}
            num_grupos = 0

        # Filas: datos + subtotales + total + resumen + contrato + firmas + espaciado
        rows_needed = len(df_reporte) + num_grupos + 1 + len(resumen_counts) + 25
        start_row = 8
        ws.insert_rows(start_row, amount=rows_needed)

        # Escribir datos con cortes por actividad
        current_row = 8
        ultima_actividad = None
        conteo_actividad = 0
        total_general = 0

        for _, row in df_reporte.iterrows():
            actividad_actual = row.get('TIPO DE ACTIVIDAD', '')
            
            # Si cambia la actividad, mostrar subtotal de la anterior
            if ultima_actividad is not None and actividad_actual != ultima_actividad:
                ws.cell(row=current_row, column=1, value="ACTIVIDADES: ")
                ws.cell(row=current_row, column=2, value=conteo_actividad).font = openpyxl.styles.Font(bold=True)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
                ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                
                for c in range(1, 10):
                    style_id = base_styles[c-1]
                    if style_id is not None:
                        # For subtotals we just want to grab the border, but we can't easily merge 
                        # just the border if we only have the _style ID. Using the whole style is fine 
                        # and ensures consistency, except it might reset the font to non-bold.
                        # So let's re-apply bold font if it resets it:
                        ws.cell(row=current_row, column=c)._style = style_id
                        
                ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True)
                ws.cell(row=current_row, column=2).font = openpyxl.styles.Font(bold=True)
                current_row += 1
                conteo_actividad = 0
            
            fecha = row.get('FECHA', '')
            if hasattr(fecha, 'strftime'):
                fecha = fecha.strftime('%Y-%m-%d')
            
            fecha_atencion = row.get('FECHA ATENCIÓN', '')
            if hasattr(fecha_atencion, 'strftime'):
                fecha_atencion = fecha_atencion.strftime('%Y-%m-%d')
            
            valores = [
                actividad_actual, fecha,
                row.get('DEPENDENCIA', ''), row.get('SOLICITANTE', ''),
                row.get('TIPO DE SOLICITUD', ''), row.get('MEDIO DE SOLICITUD', ''),
                row.get('CUMPLIDO'), fecha_atencion,
                row.get('OBSERVACIONES')
            ]
            
            for col_idx, valor in enumerate(valores, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=valor)
                if col_idx <= len(base_styles):
                    style = base_styles[col_idx - 1]
                    if style is not None:
                        cell._style = style
            
            ultima_actividad = actividad_actual
            conteo_actividad += 1
            total_general += 1
            current_row += 1
            
        # Último subtotal y Gran Total
        if ultima_actividad:
            ws.cell(row=current_row, column=1, value="ACTIVIDADES: ")
            ws.cell(row=current_row, column=2, value=conteo_actividad).font = openpyxl.styles.Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            for c in range(1, 10):
                style_id = base_styles[c-1]
                if style_id is not None:
                    ws.cell(row=current_row, column=c)._style = style_id
            ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True)
            ws.cell(row=current_row, column=2).font = openpyxl.styles.Font(bold=True)
            current_row += 1
            
            ws.cell(row=current_row, column=1, value="TOTAL GENERAL").font = openpyxl.styles.Font(bold=True, size=11)
            ws.cell(row=current_row, column=2, value=total_general).font = openpyxl.styles.Font(bold=True, size=11)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=2).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
            for c in range(1, 10): 
                style_id = base_styles[c-1]
                if style_id is not None:
                    ws.cell(row=current_row, column=c)._style = style_id
            
            ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True, size=11)
            ws.cell(row=current_row, column=2).font = openpyxl.styles.Font(bold=True, size=11)
            current_row += 1

        #current_row += 1 # Espacio

        # Fecha de informe
        ws.cell(row=current_row, column=1, value="Fecha de informe:").font = openpyxl.styles.Font(bold=True)
        ws.cell(row=current_row, column=2, value=f"{mes_actual} de {anio_actual}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
        current_row += 1

        # Datos de contrato al final (Dinamico)
        if contrato_data:
            fields = [
                ("Elaborado por:", contrato_data.get('nombre', '')),
                ("CONTRATISTA", ""),
                ("Vo.Bo:", contrato_data.get('supervisor', '')),
                ("SUPERVISOR", ""),
            ]
            for label, value in fields:
                ws.cell(row=current_row, column=1, value=label).font = openpyxl.styles.Font(bold=True)
                if value:
                    ws.cell(row=current_row, column=2, value=str(value).upper())
                    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=10)
                current_row += 1

        try:
            wb.save(output_path)
            logger.info(f"Informe guardado exitosamente en: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error al guardar el archivo de salida {output_path}: {e}")
            return False
    except Exception as e:
        logger.exception(f"Excepción no controlada en generar_informe_template: {e}")
        return False

@medir_tiempo
def analizar_plantilla_contrato(df=None, contrato_data=None, output_json_path=None):
    try:
        import openpyxl, json
        from config import TEMPLATE_EXCEL, DATA_DIR
        if contrato_data is None:
            try:
                from database import obtener_configuracion_usuario
                cfg = obtener_configuracion_usuario('admin') or {}
                contrato_data = cfg.get('datos_contrato', {}) if isinstance(cfg, dict) else {}
            except Exception:
                contrato_data = {}
        if df is None:
            df = cargar_registros(None)
        contrato_values = {
            'NRO_CONTRATO': (contrato_data.get('nro') or ''),
            'OBJETO': (contrato_data.get('objeto') or ''),
            'NOMBRE_CONTRATISTA': (contrato_data.get('nombre') or ''),
            'CEDULA': (contrato_data.get('cedula') or ''),
            'SUPERVISOR': (contrato_data.get('supervisor') or ''),
            'RANGO_FECHAS': ''
        }
        try:
            if not df.empty and 'FECHA' in df.columns:
                fechas_dt = pd.to_datetime(df['FECHA'], errors='coerce').dropna()
                if not fechas_dt.empty:
                    contrato_values['RANGO_FECHAS'] = f"{fechas_dt.min().strftime('%d/%m/%Y')} al {fechas_dt.max().strftime('%d/%m/%Y')}"
        except Exception:
            pass
        wb = openpyxl.load_workbook(TEMPLATE_EXCEL)
        ws = wb.active
        def _normalize(s):
            s = str(s).strip().upper().replace(':','')
            s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
            return ''.join(s.split())
        label_map = {
            'NRO_CONTRATO': [
                "NO.CONVENIOOCONTRATO","NOCONVENIOOCONTRATO","N°CONVENIOOCONTRATO","NROCONVENIOOCONTRATO",
                "NO.CONTRATO","NOCONTRATO","NROCONTRATO","NºCONTRATO"
            ],
            'OBJETO': ["OBJETODELCONTRATO","OBJETO"],
            'NOMBRE_CONTRATISTA': ["NOMBREDELCONTRATISTA","NOMBRECONTRATISTA"],
            'CEDULA': ["NODEIDENTIFICACIÓN","NODEIDENTIFICACION","CEDULA","CÉDULA","NIT"],
            'RANGO_FECHAS': ["FECHADEACTIVIDADES","RANGODEFECHAS"]
        }
        found = {}
        max_r = min(ws.max_row, 30)
        max_c = min(ws.max_column, 15)
        for r in range(1, max_r+1):
            for c in range(1, max_c+1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, str):
                    norm = _normalize(cell.value)
                    for key, variants in label_map.items():
                        for v in variants:
                            if v in norm and key not in found:
                                found[key] = {'label_cell': {'row': r, 'col': c}, 'value_cell': {'row': r, 'col': c+1}, 'expected_value': contrato_values.get(key, '')}
                                break
        report = {'template': TEMPLATE_EXCEL, 'found': found, 'values': contrato_values}
        if not output_json_path:
            output_json_path = os.path.join(DATA_DIR, 'analisis_plantilla.json')
        try:
            with open(output_json_path, 'w', encoding='utf-8') as f:
                json.dump(report, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        return report
    except Exception as e:
        from config import logger
        logger.error(f"Error analizando plantilla: {e}")
        return {}
