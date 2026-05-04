import openpyxl
import os

def diagnose_output():
    output_file = "test_final_report_output_v2.xlsx"
    if not os.path.exists(output_file):
        print(f"Error: Output not found at {output_file}")
        return

    wb = openpyxl.load_workbook(output_file, data_only=True)
    ws = wb.active
    
    print(f"--- Diagnóstico de SALIDA: {output_file} ---")
    # Verificando filas 1 a 15 de la cabecera y pie
    for r in range(1, 15):
        row_str = ""
        for c in range(1, 6): # Columnas A a E
            val = ws.cell(row=r, column=c).value
            if val:
                row_str += f"[{ws.cell(row=r, column=c).coordinate}]='{val}' "
        if row_str:
            print(row_str)
            
    # También el pie de página (puede estar más abajo si hay muchos datos)
    print("\n--- Pie de Página (Filas finales si aplica) ---")
    max_r = ws.max_row
    for r in range(max_r - 5, max_r + 1):
        row_str = ""
        for c in range(1, 6):
            val = ws.cell(row=r, column=c).value
            if val:
                row_str += f"[{ws.cell(row=r, column=c).coordinate}]='{val}' "
        if row_str:
            print(row_str)
                
    print("--- Fin de Diagnóstico ---")

if __name__ == "__main__":
    diagnose_output()
