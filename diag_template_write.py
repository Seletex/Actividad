import sys
import os
import pandas as pd
import openpyxl
from datetime import datetime

# Agregar el directorio actual y Actividades al path
sys.path.append(os.getcwd())
sys.path.append(os.path.join(os.getcwd(), 'Actividades'))

try:
    from config import TEMPLATE_EXCEL, logger
    from export_service import generar_informe_template, exportar_registros_filtrados
    
    print(f"Plantilla: {TEMPLATE_EXCEL}")
    
    # Inspeccionar la plantilla
    wb_temp = openpyxl.load_workbook(TEMPLATE_EXCEL)
    print(f"Hojas en la plantilla: {wb_temp.sheetnames}")
    print(f"Hoja activa: {wb_temp.active.title}")
    
    # Intentar cargar datos reales del usuario 'admin' (o el primero que tenga)
    df = pd.read_excel('actividades.xlsx', engine='openpyxl')
    print(f"Total registros en actividades.xlsx: {len(df)}")
    
    if df.empty:
        # Crear un DF de prueba si está vacío
        print("Usando datos de prueba (DF vacío)...")
        data = {
            'ID': [1],
            'USUARIO': ['admin'],
            'TIPO DE ACTIVIDAD': ['Soporte Técnico'],
            'FECHA': [datetime.now()],
            'DEPENDENCIA': ['Sistemas'],
            'SOLICITANTE': ['Tester'],
            'TIPO DE SOLICITUD': ['Correctivo'],
            'MEDIO DE SOLICITUD': ['Email'],
            'CUMPLIDO': ['Sí'],
            'FECHA ATENCIÓN': [datetime.now().strftime('%Y-%m-%d')],
            'OBSERVACIONES': ['Prueba de escritura']
        }
        df = pd.DataFrame(data)
    else:
        print(f"Usando primer registro de la DB. Usuario: {df.iloc[0].get('USUARIO')}")
        df = df.head(1)

    test_out = "diagnostic_output.xlsx"
    print(f"Generando informe de prueba en {test_out}...")
    
    success = generar_informe_template(df, test_out)
    
    if success:
        print("✅ Función retornó True.")
        wb_res = openpyxl.load_workbook(test_out)
        ws = wb_res.active
        print(f"Contenido de celda (8,1): {ws.cell(row=8, column=1).value}")
        print(f"Contenido de celda (4,3) [Nombre]: {ws.cell(row=4, column=3).value}")
    else:
        print("❌ Función retornó False.")

except Exception as e:
    import traceback
    traceback.print_exc()
